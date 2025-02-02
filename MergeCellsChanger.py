import sys
import os.path
import openpyxl
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QVBoxLayout, QPlainTextEdit, QSpinBox, QMessageBox, \
    QDesktopWidget, QLabel, QHBoxLayout
import time


class MergeCellsChanger(QWidget):
    def __init__(self):
        super().__init__()
        self.init_ui()

    def init_ui(self):
        vbox = QVBoxLayout()

        hbox_col_bar = QHBoxLayout()
        lb_col_count = QLabel('병합할 열 개수 ', self)
        self.sb_col_count = QSpinBox(self)

        self.textarea = QPlainTextEdit(self)

        hbox_btn_bar = QHBoxLayout()
        btn_make = QPushButton(self)
        btn_clear = QPushButton(self)

        #setting Layout Box
        hbox_col_bar.addWidget(lb_col_count)
        hbox_col_bar.addWidget(self.sb_col_count)

        hbox_btn_bar.addWidget(btn_make)
        hbox_btn_bar.addWidget(btn_clear)

        vbox.addLayout(hbox_col_bar)
        vbox.addWidget(self.textarea)
        vbox.addLayout(hbox_btn_bar)

        # setting widget attribute
        self.sb_col_count.setMinimum(1)

        self.textarea.setPlaceholderText('내용 붙여넣기')
        self.textarea.setPlainText("")

        btn_make.setText('생성하기')
        btn_make.clicked.connect(self.make_excel)

        btn_clear.setText('내용 지우기')
        btn_clear.clicked.connect(lambda clear: self.textarea.clear())

        self.setLayout(vbox)
        self.setWindowTitle('QPushButton')
        self.center()
        self.show()

    # A 65 / a 97
    def make_excel(self):
        merged_col = self.sb_col_count.value()
        text = self.textarea.toPlainText().strip().split('\n')
        row_count = self.textarea.document().lineCount()    # 행 개수

        if text[0] == "":  # NULL
            QMessageBox.about(self, "오류", '내용을 입력하세요')
        else:   # Not NULL
            start_col = 1
            end_col = start_col + merged_col - 1

            # 엑셀 사용
            wb = openpyxl.Workbook()
            sheet = wb.active

            # 시트에 병합된 셀 동적 생성, 해당 셀에 값 넣기
            for i in range(0, row_count):
                sheet.merge_cells(start_row=i + 1, start_column=start_col, end_row=i + 1, end_column=end_col)
                sheet.cell(row=i + 1, column=1).value = text[i].strip()

            folder_path     = ".\\MCC_Folder\\"
            filename        = folder_path + "MCC.xlsx"  # .\\MCC_Folder\\MCC.xlsx
            tmp_filename    = folder_path + "~$MCC.xlsx"

            if os.path.isdir(folder_path):  # 경로에 폴더 있을 시
                if os.path.isfile(tmp_filename):    # 파일 실행중일 때(~$ 임시 파일이 있을 시)
                    now = self.get_time()   # 현재시간
                    filename = folder_path + "MCC_%s.xlsx" % now    # 현재 시간으로 파일명 생성
            else:   # 경로에 폴더 없을 시 폴더 생성
                os.mkdir(folder_path)

            wb.save(filename)
            QMessageBox.about(self, "성공", "파일이 생성되었습니다")

    # 파일 저장할때 날짜 형식 가져오는 함수
    def get_time(self):
        now = time.strftime('%y%m%d_%H%M%S')
        return str(now)

    # gui 중앙 위치
    def center(self):
        # geometry of the main window
        qr = self.frameGeometry()

        # center point of screen
        cp = QDesktopWidget().availableGeometry().center()

        # move rectangle's center point to screen's center point
        qr.moveCenter(cp)

        # top left of rectangle becomes top left of window centering it
        self.move(qr.topLeft())


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = MergeCellsChanger()
    sys.exit(app.exec_())