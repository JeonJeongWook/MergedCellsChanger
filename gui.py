import sys
import os.path
import openpyxl
from PyQt5.QtWidgets import QApplication, QWidget, QPushButton, QVBoxLayout, QPlainTextEdit, QSpinBox, QMessageBox, \
    QDesktopWidget
import time


class GUI(QWidget):
    def __init__(self):
        super().__init__()
        self.init_ui()

    def init_ui(self):
        self.col_count = QSpinBox(self)
        self.textarea = QPlainTextEdit(self)
        vbox = QVBoxLayout()
        btn_make = QPushButton(self)
        btn_clear = QPushButton(self)
        # 병합할 열 개수 | spinbox | 미리보기(A:A, A:B)

        vbox.addWidget(self.col_count)
        vbox.addWidget(self.textarea)
        vbox.addWidget(btn_make)
        vbox.addWidget(btn_clear)

        self.col_count.setMinimum(1)

        self.textarea.setPlaceholderText('내용 붙여넣기')
        self.textarea.setPlainText("")

        btn_make.setText('생성하기')
        btn_make.clicked.connect(self.make_excel)

        btn_clear.setText('지우기')
        # btn_clear.clicked.connect(lambda clear: self.textarea.clear())
        btn_clear.clicked.connect(self.delete_excel)

        self.setLayout(vbox)
        self.setWindowTitle('QPushButton')
        self.center()
        self.show()

    # A 65 / a 97
    def make_excel(self):
        filename = ".\\MCC.xlsx"
        tmp_filename = ".\\~$MCC.xlsx"

        merged_col = self.col_count.value()
        text = self.textarea.toPlainText().strip()
        row_count = self.textarea.document().lineCount()

        if text == "":  # NULL
            QMessageBox.about(self, "오류", '내용을 입력하세요')
        else:  # Not NULL
            start_col = 1
            end_col = start_col + merged_col - 1
            text = self.textarea.toPlainText().split('\n')  # 줄바꿈 기준으로 글자 자르기

            # 엑셀
            wb = openpyxl.Workbook()
            sheet = wb.active

            # 시트에 병합된 셀 만들기, 해당 셀에 값 넣기
            for i in range(0, row_count):
                sheet.merge_cells(start_row=i + 1, start_column=start_col, end_row=i + 1, end_column=end_col)
                sheet.cell(row=i + 1, column=1).value = text[i].strip()

            # 파일 실행중일 때(~$ 임시 파일이 있을 시) 현재 시간으로 파일명 생성
            # 파일 저장 형식[MCC_yyMMdd_hhmmss]
            if os.path.isfile(tmp_filename):
                now = self.get_time()
                filename = ".\\MCC_" + str(now) + ".xlsx"
                wb.save(filename)
            else:
                wb.save(filename)
            QMessageBox.about(self, "성공", "파일이 생성되었습니다")

    # MCC파일 삭제하는 함수
    def delete_excel(self):
        print('delete_excel 실행')

    # 파일 저장할때 날짜 형식 가져오는 함수
    def get_time(self):
        now = time.strftime('%y%m%d_%H%M%S')
        return now

    # gui 중앙 위치
    def center(self):
        # geometry of the main window
        qr = self.frameGeometry()

        # center point of screen
        cp = QDesktopWidget().availableGeometry().center()

        # move rectangle's center point to screen's center point
        qr.moveCenter(cp)

        # top left of rectangle becomes top left of window centering it
        self.move(qr.topLeft())


if __name__ == '__main__':
    app = QApplication(sys.argv)
    ex = GUI()
    sys.exit(app.exec_())

'''
    # pip install openpyxl
    # pip install PyQt5
    # https://medium.com/swlh/working-with-spreadsheets-using-python-903202509407
'''
